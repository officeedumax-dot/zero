from odoo import models, fields, api
from odoo.exceptions import ValidationError
import base64
import csv
from io import StringIO, BytesIO


class ProjectDevizExportWizard(models.TransientModel):
    _name = 'project.deviz.export.wizard'
    _description = 'Export Deviz Proiect'

    project_id = fields.Many2one(
        'project.funding',
        string="Proiect",
        required=True,
        readonly=True,
    )
    file_data = fields.Binary(string="Fișier", readonly=True)
    file_name = fields.Char(string="Nume fișier", readonly=True)

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        if (
            self.env.context.get('active_model') == 'project.funding'
            and self.env.context.get('active_id')
        ):
            res['project_id'] = self.env.context['active_id']
        return res

    def action_export(self):
        """Exportă devizul proiectului în format XLSX (Excel)."""
        self.ensure_one()

        # import local, ca să nu blocăm modulul dacă lipsește librăria
        try:
            import xlsxwriter
        except ImportError:
            raise ValidationError(
                "Pentru export în Excel (.xlsx) este necesar pachetul 'xlsxwriter' "
                "instalat pe serverul Odoo.\n"
                "Până atunci, putem adapta exportul pe CSV."
            )

        lines = self.project_id.budget_line_ids.sorted(
            key=lambda l: (l.chapter or '', l.subchapter or '', l.id)
        )

        # Structura comună pentru export/import
        headers = [
            'chapter',
            'subchapter',
            'name',
            'chelt_elig_baza',
            'chelt_elig_tva',
            'chelt_neelig_baza',
            'chelt_neelig_tva',
            'tip_cheltuiala',
            'mysmis',
            'total_chelt_eligibile_neramb',
            'total_chelt_eligibile_aport',
        ]

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        sheet = workbook.add_worksheet('Deviz')

        # Header
        for col, field in enumerate(headers):
            sheet.write(0, col, field)

        # Linii
        row = 1
        for l in lines:
            sheet.write(row, 0, l.chapter or '')
            sheet.write(row, 1, l.subchapter or '')
            sheet.write(row, 2, l.name or '')
            sheet.write(row, 3, l.chelt_elig_baza or 0.0)
            sheet.write(row, 4, l.chelt_elig_tva or 0.0)
            sheet.write(row, 5, l.chelt_neelig_baza or 0.0)
            sheet.write(row, 6, l.chelt_neelig_tva or 0.0)
            sheet.write(row, 7, l.tip_cheltuiala or '')
            sheet.write(row, 8, l.mysmis or '')
            sheet.write(row, 9, l.total_chelt_eligibile_neramb or 0.0)
            sheet.write(row, 10, l.total_chelt_eligibile_aport or 0.0)
            row += 1

        workbook.close()
        data = output.getvalue()
        output.close()

        cod = self.project_id.cod or 'proiect'
        self.file_name = f"deviz_{cod}.xlsx"
        self.file_data = base64.b64encode(data)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'project.deviz.export.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }


class ProjectDevizImportWizard(models.TransientModel):
    _name = 'project.deviz.import.wizard'
    _description = 'Import Deviz Proiect'

    project_id = fields.Many2one(
        'project.funding',
        string="Proiect",
        required=True,
        readonly=True,
    )
    file_data = fields.Binary(string="Fișier (XLSX/CSV)", required=True)
    file_name = fields.Char(string="Nume fișier")
    confirm_override = fields.Boolean(
        string="Șterge liniile de deviz existente",
        help="Bifează această opțiune pentru a șterge liniile de deviz existente înainte de import.",
    )

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        if (
            self.env.context.get('active_model') == 'project.funding'
            and self.env.context.get('active_id')
        ):
            res['project_id'] = self.env.context['active_id']
        return res

    def _read_rows_from_file(self):
        """Citește fișierul încărcat și returnează o listă de dict-uri (rânduri)."""
        if not self.file_data:
            raise ValidationError("Încărcați un fișier pentru import.")

        data = base64.b64decode(self.file_data)
        filename = (self.file_name or "").lower()
        rows = []
        headers = []

        # 1) XLSX (Excel modern)
        if filename.endswith('.xlsx'):
            try:
                import openpyxl
            except ImportError:
                raise ValidationError(
                    "Pentru import din Excel (.xlsx) este necesar pachetul 'openpyxl' "
                    "instalat pe serverul Odoo.\n"
                    "Alternativ, salvați fișierul ca CSV (delimitat de punct și virgulă) și importați din nou."
                )

            wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
            sheet = wb.active

            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [h or '' for h in header_row]
            if not any(headers):
                raise ValidationError("Fișier XLSX invalid sau fără header pe primul rând.")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row is None or all(cell is None for cell in row):
                    continue
                vals = {
                    headers[i]: (row[i] if i < len(row) and row[i] is not None else "")
                    for i in range(len(headers))
                }
                rows.append(vals)

        # 2) XLS (Excel vechi)
        elif filename.endswith('.xls'):
            try:
                import xlrd
            except ImportError:
                raise ValidationError(
                    "Pentru import din Excel (.xls) este necesar pachetul 'xlrd' "
                    "instalat pe serverul Odoo.\n"
                    "Alternativ, salvați fișierul ca XLSX sau CSV și importați din nou."
                )

            wb = xlrd.open_workbook(file_contents=data)
            sh = wb.sheet_by_index(0)

            headers = [str(h or '') for h in sh.row_values(0)]
            if not any(headers):
                raise ValidationError("Fișier XLS invalid sau fără header pe primul rând.")

            for rx in range(1, sh.nrows):
                row_vals = sh.row_values(rx)
                if not any(row_vals):
                    continue
                vals = {}
                for idx, header in enumerate(headers):
                    vals[header] = row_vals[idx] if idx < len(row_vals) else ""
                rows.append(vals)

        # 3) CSV (fallback)
        else:
            try:
                text = data.decode('utf-8-sig')
            except UnicodeDecodeError:
                raise ValidationError(
                    "Fișierul nu pare a fi CSV UTF-8.\n"
                    "Pentru import din Excel, folosiți XLSX sau CSV (UTF-8, delimitat de ';')."
                )

            reader = csv.DictReader(StringIO(text), delimiter=';')
            if not reader.fieldnames:
                raise ValidationError("Fișier CSV invalid sau fără header pe primul rând.")
            headers = reader.fieldnames
            rows = list(reader)

        # Validăm coloanele obligatorii
        required_cols = ['chapter', 'subchapter', 'name']
        header_set = set(headers or (rows[0].keys() if rows else []))

        for col in required_cols:
            if col not in header_set:
                raise ValidationError(
                    f"Coloana obligatorie «{col}» lipsește din fișier.\n"
                    f"Header așteptat: {', '.join(required_cols)}."
                )

        return rows

    def action_import(self):
        self.ensure_one()
        project = self.project_id

        if project.budget_line_ids and not self.confirm_override:
            raise ValidationError(
                "Proiectul are deja linii de deviz.\n"
                "Bifează opțiunea «Șterge liniile de deviz existente» pentru a continua importul."
            )

        rows = self._read_rows_from_file()

        # Helper pentru a converti orice valoare în string "curat"
        def _s(val):
            if val is None:
                return ''
            if isinstance(val, (int, float)):
                text = str(val)
                # opțional: dacă e 1.0, îl facem "1"
                if text.endswith('.0'):
                    text = text[:-2]
                return text.strip()
            return str(val).strip()

        # Verificăm dubluri în fișier (același capitol + subcapitol)
        seen = set()
        cleaned_rows = []
        for row in rows:
            ch = _s(row.get('chapter'))
            sub = _s(row.get('subchapter'))
            key = (ch, sub)
            if key in seen:
                raise ValidationError(
                    f"Fișierul conține linii duplicate pentru capitol «{ch}» și subcapitol «{sub}»."
                )
            seen.add(key)
            cleaned_rows.append(row)

        # Ștergem liniile existente dacă e cazul
        if project.budget_line_ids:
            project.budget_line_ids.unlink()

        BudgetLine = self.env['project.budget']

        def _f(val):
            if val is None:
                return 0.0
            if isinstance(val, (int, float)):
                return float(val)
            val = str(val).strip()
            if not val:
                return 0.0
            # acceptăm atât 1234.56 cât și 1234,56
            return float(val.replace(',', '.'))

        # Creăm noile linii
        for row in cleaned_rows:
            vals = {
                'project_id': project.id,
                'chapter': _s(row.get('chapter')),
                'subchapter': _s(row.get('subchapter')),
                'name': _s(row.get('name')),
                'chelt_elig_baza': _f(row.get('chelt_elig_baza')),
                'chelt_elig_tva': _f(row.get('chelt_elig_tva')),
                'chelt_neelig_baza': _f(row.get('chelt_neelig_baza')),
                'chelt_neelig_tva': _f(row.get('chelt_neelig_tva')),
                'tip_cheltuiala': _s(row.get('tip_cheltuiala')),
                'mysmis': _s(row.get('mysmis')),
                'total_chelt_eligibile_neramb': _f(row.get('total_chelt_eligibile_neramb')),
                'total_chelt_eligibile_aport': _f(row.get('total_chelt_eligibile_aport')),
            }
            BudgetLine.create(vals)

        # Revenim pe proiect
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'project.funding',
            'view_mode': 'form',
            'res_id': project.id,
            'target': 'current',
        }
